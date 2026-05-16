"""
Router de reportes, estadísticas y exportación.
Endpoints:
  GET /reportes/dashboard-stats      → todos los KPIs para el dashboard
  GET /reportes/exportar/{tipo}      → CSV/Excel de cualquier capa
  GET /reportes/exportar/simulacion/{id} → CSV/Excel de resultados de simulación
"""
import io
import csv
from typing import Literal
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.auth.dependencies import get_current_user
from app.models.red import Tuberia, Nodo, Valvula, Tanque, Fuente, Dano
from app.models.simulacion import Simulacion

router = APIRouter(prefix="/reportes", tags=["Reportes y exportación"])
CanView = Depends(get_current_user)

# ── Paleta de estados para demo data ─────────────────────────
DEMO_STATS = {
    "es_demo": True,
    "red": {
        "total_tuberias": 42,
        "total_nodos": 35,
        "total_valvulas": 8,
        "total_tanques": 2,
        "total_fuentes": 1,
        "km_red": 3.8,
    },
    "estados_tuberias": {"Bueno": 25, "Regular": 10, "Malo": 5, "Critico": 2},
    "materiales_tuberias": {"PVC": 22, "AC": 12, "HF": 6, "PE": 2},
    "diametros_tuberias": {"25mm": 5, "50mm": 18, "75mm": 12, "100mm": 5, "150mm": 2},
    "ultima_simulacion": {
        "id": None, "nombre": "Demo Labateca",
        "presion_min": 14.2, "presion_max": 34.8,
        "presion_media": 24.5, "nodos_criticos": 1,
        "nodos_total": 6, "tuberias_total": 7,
    },
    "historial_simulaciones": [],
    "presiones_nodos": {
        "NOD-001": 28.5, "NOD-002": 24.3, "NOD-003": 19.8,
        "NOD-004": 22.1, "NOD-005": 17.4, "NOD-006": 14.2,
    },
    "usuarios": {
        "total_usuarios": 1520,
        "demanda_total_lps": 12.5,
        "por_tipo": {"Residencial": 1400, "Comercial": 100, "Institucional": 20}
    }
}


# ══════════════════════════════════════════════════════════════
# DASHBOARD STATS
# ══════════════════════════════════════════════════════════════
@router.get("/dashboard-stats")
async def dashboard_stats(db: AsyncSession = Depends(get_db), _=CanView):
    """
    Retorna todos los KPIs, distribuciones y datos de series para el dashboard.
    Si la BD está vacía, retorna datos de demostración.
    """
    async def count(model):
        return (await db.execute(select(func.count()).select_from(model))).scalar() or 0

    total_tub = await count(Tuberia)
    total_nod = await count(Nodo)

    if total_tub == 0 and total_nod == 0:
        return DEMO_STATS

    # ── Conteos generales ────────────────────────────────────
    total_val = await count(Valvula)
    total_tan = await count(Tanque)
    total_fue = await count(Fuente)
    total_danos = await count(Dano)

    # ── Distribución por estado ──────────────────────────────
    estado_res = await db.execute(
        select(Tuberia.estado, func.count().label("n"))
        .group_by(Tuberia.estado).order_by(func.count().desc())
    )
    estados = {row.estado or "Desconocido": row.n for row in estado_res}

    # ── Distribución por material ────────────────────────────
    mat_res = await db.execute(
        select(Tuberia.material, func.count().label("n"))
        .group_by(Tuberia.material).order_by(func.count().desc())
    )
    materiales = {row.material or "Sin dato": row.n for row in mat_res}

    # ── Distribución por diámetro ────────────────────────────
    diam_res = await db.execute(
        select(Tuberia.diametro_mm, func.count().label("n"))
        .group_by(Tuberia.diametro_mm).order_by(Tuberia.diametro_mm)
    )
    diametros = {f"{int(row.diametro_mm)}mm": row.n for row in diam_res if row.diametro_mm}

    # ── Última simulación ────────────────────────────────────
    sim_res = await db.execute(
        select(Simulacion).where(Simulacion.estado == "completada")
        .order_by(Simulacion.fecha_creacion.desc()).limit(1)
    )
    ultima_sim = sim_res.scalar_one_or_none()

    ultima_sim_data = None
    presiones_nodos = {}

    if ultima_sim:
        ultima_sim_data = {
            "id": ultima_sim.id, "nombre": ultima_sim.nombre,
            "presion_min": ultima_sim.presion_min_mca,
            "presion_max": ultima_sim.presion_max_mca,
            "presion_media": ultima_sim.presion_media_mca,
            "nodos_criticos": ultima_sim.nodos_criticos,
            "nodos_total": ultima_sim.nodos_total,
            "tuberias_total": ultima_sim.tuberias_total,
        }
        if ultima_sim.resultados_nodos:
            presiones_nodos = {
                k: v.get("presion_mca", 0)
                for k, v in ultima_sim.resultados_nodos.items()
            }

    # ── Historial de simulaciones ────────────────────────────
    hist_res = await db.execute(
        select(Simulacion).where(Simulacion.estado == "completada")
        .order_by(Simulacion.fecha_creacion).limit(20)
    )
    historial = [
        {
            "nombre": s.nombre,
            "fecha": s.fecha_creacion.isoformat() if s.fecha_creacion else None,
            "presion_min": s.presion_min_mca, "presion_max": s.presion_max_mca,
            "presion_media": s.presion_media_mca, "nodos_criticos": s.nodos_criticos,
        }
        for s in hist_res.scalars().all()
    ]

    # ── Consumos y usuarios ──────────────────────────────────
    usuarios_res = await db.execute(
        select(
            func.sum(Nodo.num_usuarios).label("total_usuarios"),
            func.sum(Nodo.demanda_base_lps).label("demanda_total_lps")
        )
    )
    u_row = usuarios_res.fetchone()
    total_usuarios = int(u_row.total_usuarios or 0)
    demanda_total_lps = float(u_row.demanda_total_lps or 0.0)

    tipo_usu_res = await db.execute(
        select(Nodo.tipo_usuario, func.sum(Nodo.num_usuarios).label("n"))
        .where(Nodo.tipo_usuario.is_not(None))
        .group_by(Nodo.tipo_usuario)
    )
    usuarios_por_tipo = {row.tipo_usuario: int(row.n or 0) for row in tipo_usu_res if row.tipo_usuario}

    return {
        "es_demo": False,
        "red": {
            "total_tuberias": total_tub, "total_nodos": total_nod,
            "total_valvulas": total_val, "total_tanques": total_tan,
            "total_fuentes": total_fue, "total_danos": total_danos, "km_red": None,
        },
        "estados_tuberias": estados,
        "materiales_tuberias": materiales,
        "diametros_tuberias": diametros,
        "ultima_simulacion": ultima_sim_data,
        "historial_simulaciones": historial,
        "presiones_nodos": presiones_nodos,
        "usuarios": {
            "total_usuarios": total_usuarios,
            "demanda_total_lps": round(demanda_total_lps, 2),
            "por_tipo": usuarios_por_tipo
        }
    }


# ══════════════════════════════════════════════════════════════
# EXPORTACIÓN DE CAPAS
# ══════════════════════════════════════════════════════════════
EXPORTABLE = {
    "tuberias": (Tuberia, ["codigo","diametro_mm","material","rugosidad_hw","year_instalacion","estado","zona_presion","sector","observaciones"]),
    "nodos":    (Nodo,    ["codigo","tipo","cota_msnm","demanda_base_lps","tipo_usuario","num_usuarios","estado","observaciones"]),
    "valvulas": (Valvula, ["codigo","tipo","estado","diametro_mm","cota_msnm","presion_setting","observaciones"]),
    "tanques":  (Tanque,  ["codigo","nombre","cota_fondo_msnm","cota_techo_msnm","capacidad_m3","nivel_max_m","material","estado"]),
    "fuentes":  (Fuente,  ["codigo","nombre","tipo","cota_piezometrica_msnm","caudal_disponible_lps","estado"]),
    "danos":    (Dano,    ["codigo","tipo_dano","severidad","estado_reparacion","costo_reparacion","volumen_perdido_est_m3","fecha_reporte","observaciones"]),
}


@router.get("/exportar/{tipo}")
async def exportar_catastro(
    tipo: Literal["tuberias","nodos","valvulas","tanques","fuentes","danos"],
    formato: Literal["csv","excel"] = Query("csv"),
    db: AsyncSession = Depends(get_db),
    _=CanView,
):
    """Exporta la capa catastral en CSV o Excel."""
    if tipo not in EXPORTABLE:
        raise HTTPException(400, "Tipo de capa no válido")

    model_cls, fields = EXPORTABLE[tipo]
    result = await db.execute(select(model_cls))
    rows = result.scalars().all()

    data = [{f: getattr(r, f, None) for f in fields} for r in rows]
    filename = f"catastro_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M')}"

    if formato == "csv":
        return _csv_response(fields, data, filename)
    return _excel_response(fields, data, filename, sheet_name=tipo.capitalize())


@router.get("/exportar/simulacion/{sim_id}")
async def exportar_simulacion(
    sim_id: int,
    formato: Literal["csv","excel"] = Query("csv"),
    db: AsyncSession = Depends(get_db),
    _=CanView,
):
    """Exporta resultados de una simulación (nodos + tuberías)."""
    result = await db.execute(select(Simulacion).where(Simulacion.id == sim_id))
    sim = result.scalar_one_or_none()
    if not sim:
        raise HTTPException(404, "Simulación no encontrada")

    filename = f"simulacion_{sim.nombre.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}"

    node_fields = ["codigo", "presion_mca", "cota_piezometrica"]
    node_data = [{"codigo": k, **v} for k, v in (sim.resultados_nodos or {}).items()]

    pipe_fields = ["codigo", "velocidad_ms", "caudal_lps"]
    pipe_data = [{"codigo": k, **v} for k, v in (sim.resultados_tuberias or {}).items()]

    if formato == "csv":
        return _csv_response(node_fields, node_data, filename + "_nodos")

    # Excel con dos hojas
    wb = openpyxl.Workbook()
    _fill_sheet(wb.active, "Nodos", node_fields, node_data)
    _fill_sheet(wb.create_sheet("Tuberías"), "Tuberías", pipe_fields, pipe_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


# ── Helpers ──────────────────────────────────────────────────
def _csv_response(fields: list, data: list, filename: str) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}.csv"'}
    return StreamingResponse(io.BytesIO(buf.getvalue().encode("utf-8")), media_type="text/csv", headers=headers)


def _excel_response(fields: list, data: list, filename: str, sheet_name: str = "Datos") -> StreamingResponse:
    wb = openpyxl.Workbook()
    _fill_sheet(wb.active, sheet_name, fields, data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}.xlsx"'}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


def _fill_sheet(ws, title: str, fields: list, data: list):
    """Rellena una hoja de Excel con cabecera azul y datos."""
    ws.title = title
    header_fill = PatternFill("solid", fgColor="0E4C92")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(fields)
    for col_idx, _ in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for row in data:
        ws.append([row.get(f) for f in fields])
